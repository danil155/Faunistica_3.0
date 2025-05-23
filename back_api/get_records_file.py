from fastapi import APIRouter, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import io
from datetime import datetime
import logging

from database.database import get_session
from database.crud import get_user_records
from .token import get_current_user
from .rate_limiter import limiter

logger = logging.getLogger(__name__)
router = APIRouter()

COLUMN_MAPPING = {
    'datetime': 'Дата добавления записи',
    'adm_country': 'Страна',
    'adm_region': 'Регион',
    'adm_district': 'Район',
    'adm_loc': 'Место сбора',
    'geo_nn': 'Широта (десятич.)',
    'geo_ee': 'Долгота (десятич.)',
    'geo_nn_raw': 'Широта (изнач.)',
    'geo_ee_raw': 'Долгота (изнач.)',
    'geo_origin': 'Происхождение координат',
    'geo_REM': 'Примечания к расположению',
    'eve_YY': 'Год',
    'eve_MM': 'Месяц',
    'eve_DD': 'День',
    'eve_day_def': 'Определён ли день',
    'eve_habitat': 'Биотоп',
    'eve_effort': 'Выборочное усиление',
    'abu_coll': 'Коллектор',
    'eve_REM': 'Примечания к сбору материала',
    'tax_fam': 'Семейство',
    'tax_gen': 'Род',
    'tax_sp': 'Вид',
    'tax_sp_def': 'Определён ли вид',
    'tax_nsp': 'Описан ли как новый вид',
    'type_status': 'Типовой статус',
    'tax_REM': 'Таксономические примечания',
    'abu': 'Общее кол-во особей',
    'abu_details': 'Кол-во особей каждого пола/зрелости',
    'abu_ind_rem': 'Комментарий к особям',
    'geo_uncert': 'Радиус неточности координат, м',
    'eve_YY_end': 'Конечный год',
    'eve_MM_end': 'Конечный месяц',
    'eve_DD_end': 'Конечный день',
}


@router.post("/get_records_data")
@limiter.limit("1/minute")
async def get_records_data(
        request: Request,
        user_data: dict = Depends(get_current_user),
        session: AsyncSession = Depends(get_session)
):
    user_id = int(user_data["sub"])
    username = user_data["username"]
    try:
        records = await get_user_records(session, user_id)

        if not records:
            logger.warning(f'No records found for user: {username} - {user_id}')
            raise HTTPException(status_code=404, detail="No records found for this user")

        def generate_excel():
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active

            headers = [
                COLUMN_MAPPING[field]
                for field in COLUMN_MAPPING.keys()
                if field not in ['id', 'publ_id', 'ip', 'errors', 'type', 'adm_verbatim']
            ]
            ws.append(headers)

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
                ws.cell(row=1, column=col).font = Font(bold=True)

            for record in records:
                row = [
                    getattr(record, field)
                    for field in COLUMN_MAPPING.keys()
                    if field not in ['id', 'publ_id', 'ip', 'errors', 'type', 'adm_verbatim']
                ]
                ws.append(row)

            wb.save(output)
            output.seek(0)
            yield output.read()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        headers = {
            "Content-Disposition": f"attachment; filename=user_{username}_records_{timestamp}.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }

        return StreamingResponse(
            generate_excel(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        logger.error(f'Exception in get_records_data: {str(e)}', exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
